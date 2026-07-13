#!/usr/bin/env python3
# Copyright (c) 2026 Contributors to the RISCV UnifiedDB <https://github.com/riscv/riscv-unified-db>
# SPDX-License-Identifier: BSD-3-Clause-Clear
"""
Phase 7: Generate the final parameter spreadsheet.

Consolidates the V2 deduplicated LLM extraction results (Phase 6) with the
UDB ground truth (Phase 1) and the LLM↔UDB alignment (Phase 5) into a single
authoritative spreadsheet covering every confirmed parameter — both existing
UDB parameters and newly discovered ones.

Inputs (defaults; configurable via CLI):
  - param_extraction/results/v2/deduped_claude-sonnet-4.json   (Phase 6)
  - param_extraction/results/v2/alignment_claude-sonnet-4.json (Phase 5/6)
  - param_extraction/data/ground_truth.json                    (Phase 1)
  - param_extraction/data/udb_param_names.txt                  (Phase 1)

Outputs:
  - param_extraction/data/parameters.csv   (CSV for programmatic use)
  - param_extraction/data/parameters.xlsx  (Excel for review / presentation)
  - param_extraction/data/parameters_stats.txt (human-readable summary)

Columns (per Phase 7 acceptance criteria):
  adoc_file, line_number, excerpt, parameter_name, named,
  class, value_type, confidence, notes
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import re
import sys
from collections import Counter
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR = PROJECT_DIR / "data"
RESULTS_DIR = PROJECT_DIR / "results"

logger = logging.getLogger("phase7")

CONFIDENCE_RANK = {"high": 3, "medium": 2, "low": 1}

COLUMNS = [
    "adoc_file",
    "line_number",
    "excerpt",
    "parameter_name",
    "named",
    "class",
    "value_type",
    "confidence",
    "notes",
]


# ── Loaders ────────────────────────────────────────────────────────────────


def load_json(path: Path) -> dict:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def load_udb_names(path: Path) -> set[str]:
    with open(path, encoding="utf-8") as f:
        return {line.strip() for line in f if line.strip()}


# ── Name normalization for newly discovered parameters ─────────────────────

# Prefixes the UDB naming convention reserves for specific intents (per Phase 7
# issue). When the LLM produces a new (non-UDB) parameter, we normalize it to
# match the existing style.

REPORT_PREFIX_PATTERNS = [
    (re.compile(r"^REPORT[_\s]", re.I), "REPORT_"),
    (re.compile(r"^TRAP_ON[_\s]", re.I), "TRAP_ON_"),
]


def normalize_name(name: str) -> str:
    """Coerce a candidate parameter name to UDB ALL_CAPS_WITH_UNDERSCORES style."""
    s = (name or "").strip()
    if not s:
        return s
    # Strip backticks / quotes if the LLM ever wrapped the name
    s = s.strip("`'\"")
    # Replace non-identifier separators with underscores
    s = re.sub(r"[^A-Za-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s.upper()


# ── Alignment lookup ───────────────────────────────────────────────────────


def build_alignment_lookup(alignment: dict) -> dict[str, list[dict]]:
    """Map llm_name -> all alignment entries for that LLM finding."""
    lookup: dict[str, list[dict]] = {}
    for a in alignment.get("alignments", []):
        llm = a.get("llm_name")
        if llm:
            lookup.setdefault(llm, []).append(a)
    return lookup


def resolved_name_and_named(
    llm_name: str,
    alignment_lookup: dict[str, list[dict]],
    udb_names: set[str],
) -> tuple[str, str, str]:
    """Return (final_name, named_flag, notes_hint)."""
    raw_llm = (llm_name or "").strip()
    normalized = normalize_name(raw_llm)

    if normalized in udb_names:
        return normalized, "yes", ""

    aligns = alignment_lookup.get(raw_llm) or alignment_lookup.get(normalized) or []
    non_none_aligns = [a for a in aligns if a.get("udb_name") and a.get("match_type") != "none"]

    if len(non_none_aligns) == 1:
        align = non_none_aligns[0]
        match_type = align["match_type"]
        udb = align["udb_name"]
        if match_type == "exact":
            return udb, "yes", ""
        hint = f"aligned via {match_type} to UDB {udb}"
        return udb, "yes", hint

    if len(non_none_aligns) > 1:
        names = ", ".join(sorted({a["udb_name"] for a in non_none_aligns}))
        return normalized, "no", f"needs review: multiple UDB matches ({names})"

    return normalized, "no", "newly discovered"


# ── Core build ─────────────────────────────────────────────────────────────


def build_rows(
    deduped: dict,
    alignment: dict,
    udb_names: set[str],
    min_confidence: str = "medium",
) -> list[dict]:
    """Build one spreadsheet row per confirmed parameter."""
    min_rank = CONFIDENCE_RANK[min_confidence]
    alignment_lookup = build_alignment_lookup(alignment)

    rows: list[dict] = []
    skipped_low = 0
    skipped_missing_excerpt = 0

    for p in deduped.get("parameters", []):
        conf = (p.get("confidence") or "low").lower()
        if CONFIDENCE_RANK.get(conf, 0) < min_rank:
            skipped_low += 1
            continue

        excerpt = (p.get("excerpt") or "").strip()
        if not excerpt:
            skipped_missing_excerpt += 1
            continue

        llm_name = p.get("parameter_name") or ""
        # Prefer existing_udb_name when the LLM itself supplied one (Phase 4
        # output sometimes includes the match) — falls through to alignment
        # lookup otherwise.
        existing = p.get("existing_udb_name")
        if existing and existing in udb_names:
            final_name, named_flag, notes_hint = existing, "yes", ""
        else:
            final_name, named_flag, notes_hint = resolved_name_and_named(
                llm_name, alignment_lookup, udb_names
            )

        rows.append(
            {
                "adoc_file": p.get("_source_file", ""),
                "line_number": int(p.get("line_number") or 0),
                "excerpt": excerpt,
                "parameter_name": final_name,
                "named": named_flag,
                "class": p.get("class", "UNKNOWN"),
                "value_type": p.get("value_type", ""),
                "confidence": conf,
                "notes": notes_hint,
            }
        )

    # Stable sort: file, then line number, then parameter name.
    rows.sort(key=lambda r: (r["adoc_file"], r["line_number"], r["parameter_name"]))

    logger.info(
        "Built %d rows (skipped %d low-confidence, %d missing-excerpt)",
        len(rows),
        skipped_low,
        skipped_missing_excerpt,
    )
    return rows


# ── Writers ────────────────────────────────────────────────────────────────


def write_csv(rows: list[dict], path: Path) -> None:
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        writer.writerows(rows)
    logger.info("Wrote %s (%d rows)", path, len(rows))


def write_xlsx(rows: list[dict], path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl not installed; skipping XLSX output. `uv add openpyxl`")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "parameters"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.append(COLUMNS)
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append([r[c] for c in COLUMNS])

    widths = {
        "adoc_file": 26,
        "line_number": 10,
        "excerpt": 80,
        "parameter_name": 40,
        "named": 7,
        "class": 18,
        "value_type": 12,
        "confidence": 12,
        "notes": 36,
    }
    for idx, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col, 18)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(COLUMNS)):
        for cell in row:
            cell.alignment = wrap

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    logger.info("Wrote %s (%d rows)", path, len(rows))


def write_stats(rows: list[dict], path: Path, ground_truth_count: int) -> None:
    classes = Counter(r["class"] for r in rows)
    value_types = Counter(r["value_type"] for r in rows)
    confidences = Counter(r["confidence"] for r in rows)
    files = Counter(r["adoc_file"] for r in rows)
    named_yes = sum(1 for r in rows if r["named"] == "yes")
    named_no = len(rows) - named_yes

    with open(path, "w", encoding="utf-8") as f:
        f.write("Phase 7 — Final Parameter Spreadsheet — Statistics\n")
        f.write("=" * 60 + "\n\n")
        f.write(f"Total rows               : {len(rows)}\n")
        f.write(f"UDB ground-truth params  : {ground_truth_count}\n")
        f.write(f"Already named in UDB     : {named_yes}\n")
        f.write(f"New (unnamed) discoveries: {named_no}\n")
        f.write(f"Spec files touched       : {len(files)}\n\n")

        f.write("Classification breakdown:\n")
        for cls, n in classes.most_common():
            f.write(f"  {cls:18s} {n:4d}\n")
        f.write("\nValue type breakdown:\n")
        for vt, n in value_types.most_common():
            f.write(f"  {vt:12s} {n:4d}\n")
        f.write("\nConfidence breakdown:\n")
        for conf, n in confidences.most_common():
            f.write(f"  {conf:10s} {n:4d}\n")
        f.write("\nTop spec files by parameter count:\n")
        for fname, n in files.most_common(15):
            f.write(f"  {fname:32s} {n:4d}\n")

    logger.info("Wrote %s", path)


# ── CLI ────────────────────────────────────────────────────────────────────


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument(
        "--deduped",
        type=Path,
        default=RESULTS_DIR / "v2" / "deduped_claude-sonnet-4.json",
        help="Path to deduped LLM results JSON (Phase 6 V2 output)",
    )
    p.add_argument(
        "--alignment",
        type=Path,
        default=RESULTS_DIR / "v2" / "alignment_claude-sonnet-4.json",
        help="Path to LLM↔UDB alignment JSON (Phase 5/6 output)",
    )
    p.add_argument(
        "--udb-names",
        type=Path,
        default=DATA_DIR / "udb_param_names.txt",
        help="Flat list of UDB parameter names (Phase 1 output)",
    )
    p.add_argument(
        "--ground-truth",
        type=Path,
        default=DATA_DIR / "ground_truth.json",
        help="UDB ground-truth JSON (Phase 1 output) — used only for stats",
    )
    p.add_argument(
        "--out-csv",
        type=Path,
        default=DATA_DIR / "parameters.csv",
        help="Output CSV path",
    )
    p.add_argument(
        "--out-xlsx",
        type=Path,
        default=DATA_DIR / "parameters.xlsx",
        help="Output XLSX path",
    )
    p.add_argument(
        "--out-stats",
        type=Path,
        default=DATA_DIR / "parameters_stats.txt",
        help="Output stats path",
    )
    p.add_argument(
        "--min-confidence",
        choices=["low", "medium", "high"],
        default="medium",
        help="Drop rows below this confidence level (default: medium; per Phase 8 spec)",
    )
    p.add_argument("-v", "--verbose", action="store_true")
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    if not args.deduped.exists():
        logger.error("Missing deduped results: %s", args.deduped)
        return 2
    if not args.alignment.exists():
        logger.error("Missing alignment results: %s", args.alignment)
        return 2
    if not args.udb_names.exists():
        logger.error("Missing UDB names list: %s", args.udb_names)
        return 2

    deduped = load_json(args.deduped)
    alignment = load_json(args.alignment)
    udb_names = load_udb_names(args.udb_names)
    ground_truth = load_json(args.ground_truth) if args.ground_truth.exists() else {"parameters": []}

    rows = build_rows(
        deduped=deduped,
        alignment=alignment,
        udb_names=udb_names,
        min_confidence=args.min_confidence,
    )

    args.out_csv.parent.mkdir(parents=True, exist_ok=True)
    write_csv(rows, args.out_csv)
    write_xlsx(rows, args.out_xlsx)
    write_stats(rows, args.out_stats, ground_truth_count=len(ground_truth.get("parameters", [])))

    return 0


if __name__ == "__main__":
    sys.exit(main())
